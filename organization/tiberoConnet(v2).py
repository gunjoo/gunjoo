import datetime
import pyodbc
import pandas as pd
import openpyxl as xl
from IPython.display import display


# CRE_DAT,  -- 생성일자
# EMP_CD,   -- 사원코드
# EMP_NM,  --사원명
# DEPT_CD,  --부서코드
# DEPT_NM,  --부서명
# POS_CD,   --직급코드
# POS_NM,   --직급명
# JSSFC_CD,  -- 직종코드
# JSSFC_NM,  -- 직종명
# PSIT_CD,   --직위코드
# PSIT_NM,   --직위명
# DUTY_CD,  --직책코드
# DUTY_NM,  --직책명
# DTY_CD,  --직무코드
# DTY_NM,  --직무명
# JBLN_CD, --직렬코드
# JBLN_NM, --직렬명
# SRCLS_CD, --호봉코드
# SRCLS_NM,  --호봉
# ECNY_DE   -- 임용일)


class SMSSender:
    def __init__(self):
        self.dnsName = "Tibero6"  # ODBC 데이터 원본 관리자(32bit) DSN
        self.dbUser = "ALLSPLUS"  # ODBC 데이터 원본 관리자(32bit) User
        self.dbPwd = "CnfAll#123"  # ODBC 데이터 원본 관리자(32bit) Password

    def connection(self):
        try:
            conn = pyodbc.connect('DSN=' + self.dnsName + ';UID=' + self.dbUser + ';PWD=' + self.dbPwd)
            cursor = conn.cursor()
            # sql = "SELECT * FROM TAB"
            # sql = "SELECT * FROM ALLSPLUS.VW_FRL_PSNT_NMPR_DTS WHERE EMP_NM='기세진'"
            # sql = """\
            #       SELECT DISTINCT EMP_CD, EMP_NM, DEPT_NM, POS_NM, PSIT_NM, DUTY_NM, DTY_NM, JBLN_NM, SRCLS_NM, ECNY_DE \
            #       FROM ALLSPLUS.VW_FRL_PSNT_NMPR_DTS\
            #       "
            sql = """ 
SELECT
	A.EMP_CD       				/* 사원코드 */
	,F.BPLC_NM       				/* 사업장 */
	,C.DEPT_NM      					/* 부서명 */
	,A.EMP_NM       				/* 사원명 */
	,D.DTS_CD_NM AS POS_CD_NM       		/* 직급 */
	,G.DTS_CD_NM AS JSSFC_NM       		/* 직종 */
	,J.DTS_CD_NM AS DTY_NM			/* 직렬 */
	,K.DTS_CD_NM AS SRCLS_CD_NM			/* 호봉 */
	,H.ECNY_DE			/* 입사일 */
	,NVL(B.RECENT_GNFD_DAT,B.ECNY_DE) AS RECENT_GNFD_DAT /* 최근발령일자 */
	,NVL(HSB.HR_GNFD_DE,B.ECNY_DE) AS HR_GNFD_DE   -- 직급(승진)일자
	,NVL(HSB.HR_GNFD_DE,B.ECNY_DE) AS HR_GNFD_DE	/* 현직급발령일 */
FROM HRI_HR_MST A
LEFT JOIN HRI_HFFC_INFO B ON A.EMP_CD = B.EMP_CD AND B.RECENT_DTA_AT = 'Y'
LEFT JOIN SYS_COMCD_DTS SCD3 ON SCD3.COM_CD = 'RSPOFC' AND SCD3.DTS_CD = B.RSPOFC /* 직책 */
LEFT JOIN SYS_COMCD_DTS SCD2 ON SCD2.COM_CD = 'S22' AND SCD2.DTS_CD = B.PSIT_CD   /* 직위 */
LEFT JOIN HRI_SALARY_BSIS HSB ON A.EMP_CD = HSB.EMP_CD AND HSB.NEWEST_INFO_AT = 'Y'
LEFT JOIN SYS_DEPT C ON B.DEPT_CD = C.DEPT_CD
LEFT JOIN SYS_COMCD_DTS D ON B.POS_CD = D.DTS_CD AND D.COM_CD = 'POS_CD'
LEFT JOIN SYS_COMCD_DTS E ON B.HFFC_SE = E.DTS_CD AND E.COM_CD = 'HFFC_SE'
LEFT JOIN SYS_BPLC F ON B.BPLC_ID = F.BPLC_ID
LEFT JOIN SYS_COMCD_DTS G ON B.JSSFC = G.DTS_CD AND G.COM_CD = 'JSSFC'
LEFT JOIN HRI_HFFC_INFO H ON A.EMP_CD = H.EMP_CD AND H.RECENT_DTA_AT = 'Y'
LEFT JOIN SYS_COMCD_DTS I ON B.PSIT_CD = I.DTS_CD AND I.COM_CD = 'PSIT_CD'
LEFT JOIN SYS_COMCD_DTS J ON B.DTY = J.DTS_CD AND J.COM_CD = 'DTY'
LEFT JOIN SYS_COMCD_DTS K ON HSB.SRCLS_CD = K.DTS_CD AND K.COM_CD = 'SRCLS_CD'
LEFT JOIN(
	SELECT EMP_CD,DUT_DES,MIN(LUP)
	FROM BSS_CHRG_MAIN_DUT
	GROUP BY EMP_CD, DUT_DES
	)BCMD ON B.EMP_CD = BCMD.EMP_CD
WHERE E.DTS_CD_NM = '재직' AND  J.DTS_CD_NM IS NOT NULL
ORDER BY C.LUP_ORD, D.LUP_ORD, NVL(HSB.HR_GNFD_DE,B.ECNY_DE), HSB.SRCLS_CD DESC, A.EMP_CD   /* 부서 , 직급 , 직급승진일 , 호봉 DESC , 사번 */
                    """
            cursor.execute(sql)
            fetchall = cursor.fetchall()
            rows = fetchall
            cursor.close()
            conn.close()
            return rows
            #
            # print("row(len)", len(rows))
            # for row in rows:
            #     print(">>>type(row)", type(row))
            #     print(">>>row", row)
        except Exception as e:
            print('>>>except:', e)
            cursor.close()
            conn.close()

    def run(self, result):
        today = datetime.datetime.now()
        # EXCEL DATA 컬럼명 추가
        emp_cd = []  # 사원코드
        emp_nm = []  # 사원명
        dept_nm = []  # 부서명
        pos_nm = []  # 직급명
        psit_nm = []  # 직위명
        duty_nm = []  # 직책명
        dty_nm = []  # 직무명
        jbln_nm = []  # 직렬명
        srcls_nm = []  # 호봉
        ecny_de = []  # 임용일
        try:
            data = result
            # ( d[2]가 부서별)
            for d in data:
                d = list(d)
                emp_cd.append(d[0])
                emp_nm.append(d[1])
                dept_nm.append(d[2])
                pos_nm.append(d[3])
                psit_nm.append(d[4])
                duty_nm.append(d[5])
                dty_nm.append(d[6])
                jbln_nm.append(d[7])
                srcls_nm.append(d[8])
                ecny_de.append(d[9])

            print('>>> 데이터프레임 가공')

            sum_df = pd.DataFrame([emp_cd, emp_nm, dept_nm, pos_nm, psit_nm, duty_nm, jbln_nm, srcls_nm, ecny_de]).T
            sum_df.columns = ['사원번호', '사원명', '부서명', '직급명', '직위명', '직책명', '직렬명', '호봉', '임용일']

            # print('>>> 작업이 완료되었습니다.')
            # filename = today.strftime('%Y%m%d') + '.csv'
            # sum_df.to_csv('2021년 조직도_'+ filename, encoding='utf-8-sig', index=False)
            # print('결과물: ', filename)

        except Exception as e:
            print('>>>except:', e)

    def data(self, result):
        print('==start date function==')
        try:
            total_data = result
            test = total_data[0]
            # 0.사원번호 1.사업장 2.부서명 3.사원명 4.직급 5.직종 6.직렬 7.호봉 8.입사일 9.최근인사발령일
            # 정책기획팀 행정6급 5호봉 정건주 \n최근발령일 :
            print('>>>test', test)
            test_d = test[2]+' '+test[6]+' '+test[4]+' '+test[7]+' '+test[3]+'\n'+'현직급입사일:'+test[8]+' '+'\n'+'최근발령일자:'+test[9]
            print('test_d:', test_d)
            # 정책기획팀 행정직 6급 정건주
            # 일반직(1호봉)
            # 현직급발령일 : 2019-09-16
            # 최근발령일자 : 2021-10-01
            entire_hstree = {}  # 전체 dict
            ceo = []  # 대표이사
            general_team = []
            policy_team = []
            financial_team = []
            supervision_team1 = []
            supervision_team2 = []
            manage_support_team = []
            ethics_team = []
            edu_team = []
            sch_edu_team = []
            area_edu_team = []
            path_search_team = []
            path_expe_team = []
            dong_ieunmteo = []
            dawon_bu_ieunmteo = []
            dawon_ma_ieunmteo = []
            song_ieunmteo = []
            mok_ieunmteo = []
            seo_bu_ieunmteo = []
            seo_ma_ieunmteo = []
            # 0.사원번호 1.사업장 2.부서명 3.사원명 4.직급 5.직종 6.직렬 7.호봉 8.입사일 9.최근인사발령일
            for d in total_data:
                # 호봉 없을 시
                if d[7]:
                    hstree = d[2]+' '+d[6]+' '+d[4]+' '+d[7]+' '+d[3]+'\n'+'현직급입사일:'+d[8]+' '+'\n'+'최근발령일자:'+d[9]
                else:
                    hstree = d[2] + ' ' + d[6] + ' ' + d[4] + ' ' + '0' + ' ' + d[3] + '\n' + '현직급입사일:' + d[8] + ' ' + '\n' + '최근발령일자:' + d[9]

                print('hstree:', hstree , 'type(hstree):', type(hstree))
                if "대표이사" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    ceo.append(hstree)
                elif "총무팀" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    general_team.append(hstree)
                elif "정책기획" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    policy_team.append(hstree)
                elif "재무회계" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    financial_team.append(hstree)
                elif "장학관" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    supervision_team1.append(hstree)
                ## 미사용
                #elif "장학관2" in hstree:
                #    hstree = ' '.join(hstree.split()[1:])
                #    supervision_team2.append(hstree)
                elif "장학사업" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    edu_team.append(hstree)
                elif "윤리경영" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    ethics_team.append(hstree)
                # 화성교육협력지원센터
                elif "운영지원" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    manage_support_team.append(hstree)
                elif "학교교육" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    sch_edu_team.append(hstree)
                elif "지역교육" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    area_edu_team.append(hstree)
                # 화성자유학년제지원센터
                elif "진로탐색" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    path_search_team.append(hstree)
                elif "진로체험" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    path_expe_team.append(hstree)
                # 동탄중앙이음터
                elif "동탄중앙" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    dong_ieunmteo.append(hstree)
                # 다원이음터센터
                elif "다원운영" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    dawon_ma_ieunmteo.append(hstree)
                elif "다원사업" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    dawon_bu_ieunmteo.append(hstree)
                # 송린이음터센터
                elif "송린" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    song_ieunmteo.append(hstree)
                # 동탄목동이음터
                elif "동탄목동" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    mok_ieunmteo.append(hstree)
                elif "서연운영" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    seo_ma_ieunmteo.append(hstree)
                elif "서연사업" in hstree:
                    hstree = ' '.join(hstree.split()[1:])
                    seo_bu_ieunmteo.append(hstree)
                else:
                    print('hstree:', hstree)
                    print('>>>예외부서 발생')

            print('ceo', ceo)
            print('general_team', general_team)
            print('policy_team', policy_team)
            print('financial_team', financial_team)

            # 부서별 key 설정
            keys = []
            re_result = []
            print('total_data', total_data)
            for d in total_data:
                re_result.append(list(d))
            for re_data in re_result:
                keys.append(re_data[2])
            team_key = list(set(keys))
            print("부서명:", team_key)
            # ['지역교육공동체팀', '장학관2팀', '송린이음터센터', '학교교육지원팀', '다원사업팀', '동탄목동이음터센터', '재무회계팀', '서연사업팀', '다원운영팀', '총무팀', '운영지원팀', '평생장학팀', '동탄목동팀', '정책기획팀', '화성교육협력지원센터', '동탄중앙팀', '송린팀', '서연운영팀', '화성시 인재육성재단', '윤리경영팀', '다원이음터센터', '진로체험팀', '장학관1팀', '진로탐색팀']
            team = []
            team.append(team_key)
            print('team:', team)

            ## make excel
            wb = xl.Workbook()
            #ws = wb.active
            ws = wb.create_sheet(title='화성시인재육성재단 조직도', index=0)
            for r in team:
                ws.append(r)
            wb.save(filename='C:/Users/USER/Desktop/jgj.xlsx')
            wb.close()

            # for row in range(team_len):
            #     cell = ws.cell(row=row, column=1)

            # for row in range(1, 4):
            #     for col in range(1, 4):
            #         cell=ws.cell(row=row, column=col)
            #         print(cell.coordinate, end="!")
            #     print()


            #tab change
            ws.sheet_properties.tabColor = "1072BA"
            # # 헤더 만들기
            # for seq, team in enumerate(team_key):
            #     ws.cell(row=1, column=seq+1, value=team)
            #
            # row_no = 2
            # for n, rows in enumerate(general_team):
            #     for seq, value in enumerate(rows):
            #         ws.cell(row=row_no+n, column=seq+1, value=value)
            # wb.save('test.xlsx')
            # wb.close()
            # print('>>>team_key', team_key)
            # print('>>>len(team_key):', len(team_key))



        except Exception as e:
            print('>>>except:', e)
            return 0


def main():
    queryResult = SMSSender().connection()
    # SMSSender().run(queryResult)
    data = SMSSender().data(queryResult)
    SMSSender().run(data)


if __name__ == '__main__':
    main()
